#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
技术部项目资料数据库(局域网版 v4)
=====================================
三层结构:项目 → BOM物料 → 资料文件

  ROOT/
    projects.json                     项目索引
    HY2026062301-浙江海港...充电仓项目/   每个项目一个文件夹(编码命名)
      总BOM清单.xlsx                   项目的 BOM Excel
      物料资料/
        8S131100AA-总进线智能电表/      每个物料一个文件夹
          图纸/xxx.pdf
          技术文档/yyy.docx

功能:
  - 登录后首页显示项目列表
  - 进入项目 → 解析 BOM Excel → 显示物料清单
  - 点物料 → 该物料全部资料(图纸/文档),可下载、可上传
  - 注册申请 + 管理员审批 + 账号管理
  - 管理员后台:操作日志(谁上传/下载了什么)

运行:  python server.py   (需 openpyxl 解析 Excel,已安装)
"""

import os
import sys
import re
import time
import json
import secrets
import threading
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import quote, unquote, urlparse, parse_qs

# ============ 配置区 ============
ROOT = r"D:\物料清单"
PORT = 8000
INIT_ADMIN = {"admin": "xxxxxxx", "name": "管理员", "dept": "系统"}
SESSION_TTL = 28800
MAX_UPLOAD = 500 * 1024 * 1024   # 上传大小上限 500MB
# ================================

HOST = "0.0.0.0"
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
AUTH_FILE = os.path.join(BASE_DIR, "auth.json")
ACTIVITY_FILE = os.path.join(BASE_DIR, "activity.json")
SESSIONS = {}
LOCK = threading.Lock()


# ==================== 数据读写 ====================

def load_auth():
    if not os.path.exists(AUTH_FILE):
        accounts = {
            "admin": {"password": INIT_ADMIN["admin"], "name": INIT_ADMIN["name"],
                      "dept": INIT_ADMIN["dept"], "role": "admin",
                      "created": time.strftime("%Y-%m-%d %H:%M")}
        }
        data = {"accounts": accounts, "applications": []}
        save_auth(data)
        return data
    with open(AUTH_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def save_auth(data):
    tmp = AUTH_FILE + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    if os.path.exists(AUTH_FILE):
        try:
            os.replace(AUTH_FILE, AUTH_FILE + ".bak")
        except OSError:
            pass
    os.replace(tmp, AUTH_FILE)


def load_projects():
    """项目索引: {编码: {name, created, bom}}"""
    pf = os.path.join(ROOT, "projects.json")
    if not os.path.exists(pf):
        return {}
    try:
        with open(pf, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_projects(data):
    pf = os.path.join(ROOT, "projects.json")
    tmp = pf + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    if os.path.exists(pf):
        try:
            os.replace(pf, pf + ".bak")
        except OSError:
            pass
    os.replace(tmp, pf)


def log_activity(user, action, project, item, filename, size=0):
    """记录操作日志(上传/下载)"""
    with LOCK:
        try:
            with open(ACTIVITY_FILE, "r", encoding="utf-8") as f:
                logs = json.load(f)
        except Exception:
            logs = []
        logs.append({
            "time": time.strftime("%Y-%m-%d %H:%M:%S"),
            "user": user, "action": action,
            "project": project, "item": item,
            "filename": filename, "size": size,
        })
        logs = logs[-2000:]  # 只保留最近 2000 条
        tmp = ACTIVITY_FILE + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(logs, f, ensure_ascii=False, indent=2)
        try:
            os.replace(tmp, ACTIVITY_FILE)
        except OSError:
            pass


# ==================== 工具函数 ====================

def fmt_size(n):
    for unit in ["B", "KB", "MB", "GB"]:
        if n < 1024:
            return f"{n:.1f} {unit}"
        n /= 1024
    return f"{n:.1f} TB"


def list_files(base):
    """递归列出目录下所有文件: [(相对路径, 绝对路径, 大小, mtime)]"""
    out = []
    if not os.path.isdir(base):
        return out
    for dirpath, dirnames, filenames in os.walk(base):
        for fn in filenames:
            full = os.path.join(dirpath, fn)
            rel = os.path.relpath(full, base)
            try:
                st = os.stat(full)
                out.append((rel, full, st.st_size, st.st_mtime))
            except OSError:
                continue
    return out


def parse_bom_excel(xlsx_path):
    """解析 BOM Excel,返回物料列表 [{code, name, spec, manufacturer, version}]。
    自动找表头列:物料号/名称/规格型号/制造商/版本号;找不到的列返回空/默认值。
    兼容旧表头(只有物料号/名称/规格/数量/单位)。
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        return None, f"解析 Excel 失败: {e}"
    if not rows:
        return [], ""

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    col = {"code": None, "name": None, "spec": None, "mfr": None, "ver": None}
    for i, h in enumerate(header):
        hl = h.lower()
        if col["code"] is None and any(k in hl for k in ("物料号", "物料编码", "物料代码", "编码", "图号", "编号", "item", "code", "part")):
            col["code"] = i
        if col["name"] is None and any(k in hl for k in ("名称", "描述", "name", "desc")) and "型号" not in hl:
            col["name"] = i
        if col["spec"] is None and any(k in hl for k in ("规格", "型号", "spec")):
            col["spec"] = i
        if col["mfr"] is None and any(k in hl for k in ("制造商", "厂商", "厂家", "品牌", "mfr", "maker")):
            col["mfr"] = i
        if col["ver"] is None and any(k in hl for k in ("版本", "版次", "version", "rev")):
            col["ver"] = i
    if col["code"] is None:
        col["code"] = 0
    if col["name"] is None:
        col["name"] = min(1, len(header) - 1)

    def cell(row, key):
        i = col.get(key)
        if i is None or i >= len(row) or row[i] is None:
            return ""
        return str(row[i]).strip()

    items = []
    for row in rows[1:]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        code = cell(row, "code")
        if code:
            items.append({
                "code": code,
                "name": cell(row, "name"),
                "spec": cell(row, "spec"),
                "manufacturer": cell(row, "mfr"),
                "version": cell(row, "ver") or "A00",
            })
    return items, ""


def bump_version(ver: str) -> str:
    """版本号递增:A00 -> A01 -> A02 ... (数字部分 +1,前缀字母保留)"""
    ver = (ver or "A00").strip()
    m = re.match(r"^(\D*)(\d+)$", ver)
    if not m:
        return "A01"
    prefix, num = m.group(1), int(m.group(2))
    width = len(m.group(2))
    return f"{prefix}{num + 1:0{width}d}"


def ensure_bom_columns(ws):
    """确保 BOM 表头有:物料号/名称/规格型号/制造商/版本号 列。
    旧表头缺列时自动追加;返回 {列名: 列号}(1 起始)。
    """
    header = [str(ws.cell(row=1, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
    def find(kws):
        for i, h in enumerate(header):
            if h and any(k in h for k in kws):
                return i + 1
        return None
    cols = {
        "code": find(["物料号", "物料编码", "物料代码", "编码", "图号", "编号", "item", "code", "part"]) or 1,
        "name": find(["名称", "描述", "name", "desc"]),
        "spec": find(["规格", "型号", "spec"]),
        "mfr": find(["制造商", "厂商", "厂家", "品牌", "mfr", "maker"]),
        "ver": find(["版本", "版次", "version", "rev"]),
    }
    # 补缺失列
    if cols["name"] is None:
        ws.cell(row=1, column=ws.max_column + 1).value = "名称"
        cols["name"] = ws.max_column
    if cols["spec"] is None:
        ws.cell(row=1, column=ws.max_column + 1).value = "规格型号"
        cols["spec"] = ws.max_column
    if cols["mfr"] is None:
        ws.cell(row=1, column=ws.max_column + 1).value = "制造商"
        cols["mfr"] = ws.max_column
    if cols["ver"] is None:
        ws.cell(row=1, column=ws.max_column + 1).value = "版本号"
        cols["ver"] = ws.max_column
    return cols


def parse_multipart(body: bytes, content_type: str):
    """解析 multipart/form-data,返回 (字段dict, 文件dict{字段名:(文件名,字节)})"""
    m = re.search(r'boundary=(?:"([^"]+)"|([^;]+))', content_type or "")
    if not m:
        return {}, {}
    boundary = (m.group(1) or m.group(2)).strip()
    delim = b"--" + boundary.encode()
    parts = body.split(delim)
    fields, files = {}, {}
    for part in parts[1:-1]:
        part = part.strip(b"\r\n")
        if not part:
            continue
        header_end = part.find(b"\r\n\r\n")
        if header_end == -1:
            continue
        headers = part[:header_end].decode("utf-8", "replace")
        content = part[header_end + 4:]
        if content.endswith(b"\r\n"):
            content = content[:-2]
        nm = re.search(r'name="([^"]+)"', headers)
        if not nm:
            continue
        name = nm.group(1)
        fnm = re.search(r'filename="([^"]*)"', headers)
        if fnm and fnm.group(1):
            fname = fnm.group(1)
            files[name] = (fname, content)
        else:
            fields[name] = content.decode("utf-8", "replace")
    return fields, files


# ==================== 页面模板 ====================

LOGIN_PAGE = """<!DOCTYPE html>
<html lang="zh">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>登录 · 技术部项目资料数据库</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: "Microsoft YaHei", "PingFang SC", sans-serif;
         background: #eef1f6; display: flex; align-items: center; justify-content: center;
         min-height: 100vh; }}
  .box {{ background: #fff; border-radius: 12px; box-shadow: 0 4px 24px rgba(0,0,0,.08);
         padding: 36px 40px; width: 380px; }}
  .logo {{ text-align: center; margin-bottom: 8px; }}
  .logo img {{ max-width: 280px; height: auto; }}
  h1 {{ font-size: 19px; text-align: center; color: #1f2329; margin-bottom: 6px; }}
  .sub {{ font-size: 13px; color: #86909c; text-align: center; margin-bottom: 24px; }}
  label {{ display: block; font-size: 13px; color: #4e5969; margin: 14px 0 6px; }}
  input {{ width: 100%; padding: 10px 12px; font-size: 14px; border: 1px solid #d9dde3;
          border-radius: 6px; outline: none; }}
  input:focus {{ border-color: #3370ff; box-shadow: 0 0 0 2px rgba(51,112,255,.12); }}
  button {{ width: 100%; margin-top: 22px; padding: 11px; font-size: 15px; color: #fff;
           background: #3370ff; border: none; border-radius: 6px; cursor: pointer; }}
  button:hover {{ background: #2a5fd6; }}
  .err {{ background: #fff3f0; color: #d54941; font-size: 13px; padding: 8px 12px;
         border-radius: 6px; margin-top: 16px; }}
  .ok {{ background: #e8f7ef; color: #23a65a; font-size: 13px; padding: 8px 12px;
        border-radius: 6px; margin-top: 16px; }}
  .link {{ text-align: center; margin-top: 18px; font-size: 13px; }}
  .link a {{ color: #3370ff; text-decoration: none; }}
  .tip {{ margin-top: 14px; font-size: 12px; color: #c0c4cc; text-align: center; }}
</style>
</head>
<body>
<div class="box">
  <div class="logo"><img src="/static/logo.png" alt="公司Logo"></div>
  <h1>技术部项目资料数据库</h1>
  <div class="sub">请输入账号密码登录</div>
  <form method="post" action="/login">
    <label>账号</label>
    <input type="text" name="user" autocomplete="username" autofocus required>
    <label>密码</label>
    <input type="password" name="pass" autocomplete="current-password" required>
    <button type="submit">登 录</button>
    {msg}
  </form>
  <div class="link">没有账号? <a href="/register">申请注册</a></div>
  <div class="tip">局域网内部系统 · 请勿外传</div>
</div>
</body>
</html>"""

REGISTER_PAGE = """<!DOCTYPE html>
<html lang="zh">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>申请注册 · 技术部项目资料数据库</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: "Microsoft YaHei", "PingFang SC", sans-serif;
         background: #eef1f6; display: flex; align-items: center; justify-content: center;
         min-height: 100vh; }}
  .box {{ background: #fff; border-radius: 12px; box-shadow: 0 4px 24px rgba(0,0,0,.08);
         padding: 36px 40px; width: 400px; }}
  .logo {{ text-align: center; margin-bottom: 8px; }}
  .logo img {{ max-width: 280px; height: auto; }}
  h1 {{ font-size: 19px; text-align: center; color: #1f2329; margin-bottom: 6px; }}
  .sub {{ font-size: 13px; color: #86909c; text-align: center; margin-bottom: 20px; }}
  label {{ display: block; font-size: 13px; color: #4e5969; margin: 12px 0 5px; }}
  input {{ width: 100%; padding: 9px 12px; font-size: 14px; border: 1px solid #d9dde3;
          border-radius: 6px; outline: none; }}
  input:focus {{ border-color: #3370ff; box-shadow: 0 0 0 2px rgba(51,112,255,.12); }}
  button {{ width: 100%; margin-top: 20px; padding: 11px; font-size: 15px; color: #fff;
           background: #3370ff; border: none; border-radius: 6px; cursor: pointer; }}
  button:hover {{ background: #2a5fd6; }}
  .msg {{ margin-top: 16px; font-size: 13px; }}
  .err {{ background: #fff3f0; color: #d54941; padding: 8px 12px; border-radius: 6px; }}
  .ok {{ background: #e8f7ef; color: #23a65a; padding: 10px 12px; border-radius: 6px; }}
  .ok a {{ color: #23a65a; font-weight: bold; }}
  .link {{ text-align: center; margin-top: 16px; font-size: 13px; }}
  .link a {{ color: #3370ff; text-decoration: none; }}
</style>
</head>
<body>
<div class="box">
  <div class="logo"><img src="/static/logo.png" alt="公司Logo"></div>
  <h1>申请注册账号</h1>
  <div class="sub">提交后需管理员审批,通过后才能登录</div>
  <form method="post" action="/register">
    <label>登录账号</label>
    <input type="text" name="user" required placeholder="如: wangwu">
    <label>密码</label>
    <input type="password" name="pass" required>
    <label>确认密码</label>
    <input type="password" name="pass2" required>
    <label>姓名</label>
    <input type="text" name="name" required placeholder="如: 王五">
    <label>部门</label>
    <input type="text" name="dept" required placeholder="如: 技术部">
    <label>申请理由(可选)</label>
    <input type="text" name="reason" placeholder="如: 需要查看 SVG 资料">
    <button type="submit">提交申请</button>
    {msg}
  </form>
  <div class="link"><a href="/login">← 返回登录</a></div>
</div>
</body>
</html>"""

HOME_PAGE = """<!DOCTYPE html>
<html lang="zh">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>项目列表 · 技术部项目资料数据库</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: "Microsoft YaHei", "PingFang SC", sans-serif;
         background: #f0f2f5; color: #1f2329; padding: 24px 16px; }}
  .wrap {{ max-width: 1080px; margin: 0 auto; }}
  .topbar {{ display: flex; justify-content: space-between; align-items: center; margin-bottom: 18px; }}
  .topbar .title {{ font-size: 20px; font-weight: bold; }}
  .topbar .user {{ font-size: 13px; color: #4e5969; }}
  .nav a {{ font-size: 13px; color: #3370ff; text-decoration: none; margin-left: 14px; }}
  .card {{ background: #fff; border-radius: 10px; box-shadow: 0 2px 12px rgba(0,0,0,.06);
          padding: 24px 28px; margin-bottom: 16px; }}
  .card h2 {{ font-size: 16px; margin-bottom: 14px; color: #1f2329; }}
  table.projlist {{ width: 100%; border-collapse: collapse; font-size: 14px; }}
  table.projlist th {{ text-align: left; color: #86909c; font-weight: 500; font-size: 12px;
       padding: 10px 12px; border-bottom: 1px solid #e5e6eb; }}
  table.projlist td {{ padding: 12px; border-bottom: 1px solid #f2f3f5; vertical-align: middle; }}
  table.projlist tr.proj {{ cursor: pointer; transition: background .12s; }}
  table.projlist tr.proj:hover {{ background: #f7f8fa; }}
  table.projlist tr.proj:hover .pname {{ color: #3370ff; }}
  .idx {{ color: #c0c4cc; font-size: 13px; text-align: center; }}
  .pname {{ font-weight: 500; line-height: 1.5; }}
  .pcode {{ font-size: 12px; color: #86909c; font-family: Consolas,monospace; margin-top: 3px; }}
  .pcreator {{ font-size: 13px; color: #4e5969; }}
  .pmeta {{ font-size: 12px; color: #86909c; }}
  .hint {{ font-size: 11px; color: #c0c4cc; white-space: nowrap; }}
  .badge {{ display: inline-block; font-size: 12px; padding: 2px 10px; border-radius: 10px; }}
  .badge-ok {{ background: #e8f7ef; color: #23a65a; }}
  .badge-no {{ background: #f4f4f5; color: #86909c; }}
  .empty {{ color: #86909c; text-align: center; padding: 30px 0; }}
  .btn {{ display: inline-block; padding: 6px 16px; font-size: 13px; border: none;
         border-radius: 5px; cursor: pointer; color: #fff; background: #3370ff;
         text-decoration: none; }}
  .btn:hover {{ opacity: .85; }}
  .btn-green {{ background: #23a65a; }}
  .btn-edit {{ background: #f5a623; }}
  .btn-edit:hover {{ background: #d98d0e; }}
  .card-head {{ display: flex; justify-content: space-between; align-items: center; margin-bottom: 14px; }}
  .card-head h2 {{ margin-bottom: 0; }}
  .footer {{ margin-top: 18px; font-size: 12px; color: #c0c4cc; text-align: center; }}
  /* ---------- 弹窗 ---------- */
  .mask {{ display: none; position: fixed; inset: 0; background: rgba(0,0,0,.45);
          z-index: 100; align-items: center; justify-content: center; }}
  .mask.show {{ display: flex; }}
  .modal {{ background: #fff; border-radius: 12px; padding: 28px 32px; width: 440px;
           box-shadow: 0 8px 40px rgba(0,0,0,.2); }}
  .modal h3 {{ font-size: 17px; margin-bottom: 18px; }}
  .modal label {{ display: block; font-size: 13px; color: #4e5969; margin: 12px 0 5px; }}
  .modal input {{ width: 100%; padding: 9px 12px; font-size: 14px; border: 1px solid #d9dde3;
                 border-radius: 6px; outline: none; }}
  .modal input:focus {{ border-color: #3370ff; box-shadow: 0 0 0 2px rgba(51,112,255,.12); }}
  .modal .btns {{ display: flex; gap: 10px; margin-top: 22px; }}
  .modal .btns button {{ flex: 1; padding: 10px; font-size: 14px; border: none;
                        border-radius: 6px; cursor: pointer; }}
  .modal .btns .ok {{ background: #3370ff; color: #fff; }}
  .modal .btns .cancel {{ background: #f2f3f5; color: #4e5969; }}
</style>
</head>
<body>
<div class="wrap">
  <div class="topbar">
    <div class="title">📁 技术部项目资料数据库</div>
    <div class="user">👤 {user} <span class="nav">{admin_nav}</span></div>
  </div>

  <div class="card">
    <div class="card-head">
      <h2>🗂 项目列表</h2>
      {newproj_btn}
    </div>
    {projects_html}
  </div>

  <div class="footer">技术部项目资料数据库 v4 · 双击项目行进入</div>
</div>

{newproj_modal}
<script>
function openModal() {{ document.getElementById('mask').classList.add('show'); }}
function closeModal() {{ document.getElementById('mask').classList.remove('show'); }}
var m = document.getElementById('mask');
if (m) m.addEventListener('click', function(e) {{
  if (e.target === this) closeModal();
}});
</script>
</body>
</html>"""

PROJECT_PAGE = """<!DOCTYPE html>
<html lang="zh">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{code} · 项目资料</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: "Microsoft YaHei", "PingFang SC", sans-serif;
         background: #f0f2f5; color: #1f2329; padding: 24px 16px; }}
  .wrap {{ max-width: 1080px; margin: 0 auto; }}
  .topbar {{ display: flex; justify-content: space-between; align-items: center; margin-bottom: 14px; }}
  .topbar .back {{ font-size: 13px; color: #3370ff; text-decoration: none; }}
  .topbar .user {{ font-size: 13px; color: #4e5969; }}
  .nav a {{ font-size: 13px; color: #3370ff; text-decoration: none; margin-left: 14px; }}
  .card {{ background: #fff; border-radius: 10px; box-shadow: 0 2px 12px rgba(0,0,0,.06);
          padding: 24px 28px; margin-bottom: 16px; }}
  .card h2 {{ font-size: 16px; margin-bottom: 14px; }}
  .pcode {{ font-size: 13px; color: #3370ff; font-family: Consolas,monospace; margin-bottom: 4px; }}
  .pname {{ font-size: 17px; font-weight: bold; margin-bottom: 6px; }}
  .pmeta {{ font-size: 12px; color: #86909c; margin-bottom: 12px; }}
  .tip {{ background: #f0f7ff; color: #2f6fed; font-size: 12px; padding: 8px 12px;
         border-radius: 6px; margin-bottom: 12px; }}
  .err {{ background: #fff3f0; color: #d54941; font-size: 13px; padding: 8px 12px;
         border-radius: 6px; margin-bottom: 12px; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 14px; }}
  th {{ text-align: left; color: #86909c; font-weight: 500; font-size: 12px;
       padding: 8px 10px; border-bottom: 1px solid #e5e6eb; }}
  td {{ padding: 9px 10px; border-bottom: 1px solid #f2f3f5; }}
  tr:hover td {{ background: #f7f8fa; }}
  .mat a {{ color: #1f2329; text-decoration: none; font-weight: 500; }}
  .mat a:hover {{ color: #3370ff; }}
  .code {{ font-family: Consolas,monospace; color: #3370ff; }}
  .cnt {{ font-size: 12px; color: #86909c; }}
  .ver {{ display: inline-block; font-size: 12px; font-weight: bold; color: #3370ff;
         background: #f0f7ff; padding: 2px 8px; border-radius: 10px; }}
  .empty {{ color: #86909c; text-align: center; padding: 24px 0; }}
  form.inline {{ display: inline; }}
  .btn {{ display: inline-block; padding: 6px 16px; font-size: 13px; border: none;
         border-radius: 5px; cursor: pointer; color: #fff; background: #3370ff;
         text-decoration: none; }}
  .btn:hover {{ opacity: .85; }}
  .btn-green {{ background: #23a65a; }}
  .btn-edit {{ background: #f5a623; }}
  .btn-edit:hover {{ background: #d98d0e; }}
  .upload {{ margin-top: 12px; padding-top: 12px; border-top: 1px solid #f2f3f5; font-size: 13px; }}
  .upload input[type=file] {{ font-size: 13px; }}
  .card-head {{ display: flex; justify-content: space-between; align-items: center; margin-bottom: 14px; }}
  .card-head h2 {{ margin-bottom: 0; }}
  /* ---------- 弹窗 ---------- */
  .mask {{ display: none; position: fixed; inset: 0; background: rgba(0,0,0,.45);
          z-index: 100; align-items: center; justify-content: center; }}
  .mask.show {{ display: flex; }}
  .modal {{ background: #fff; border-radius: 12px; padding: 28px 32px; width: 440px;
           box-shadow: 0 8px 40px rgba(0,0,0,.2); }}
  .modal h3 {{ font-size: 17px; margin-bottom: 18px; }}
  .modal label {{ display: block; font-size: 13px; color: #4e5969; margin: 12px 0 5px; }}
  .modal input {{ width: 100%; padding: 9px 12px; font-size: 14px; border: 1px solid #d9dde3;
                 border-radius: 6px; outline: none; }}
  .modal input:focus {{ border-color: #3370ff; box-shadow: 0 0 0 2px rgba(51,112,255,.12); }}
  .modal .btns {{ display: flex; gap: 10px; margin-top: 22px; }}
  .modal .btns button {{ flex: 1; padding: 10px; font-size: 14px; border: none;
                        border-radius: 6px; cursor: pointer; }}
  .modal .btns .ok {{ background: #3370ff; color: #fff; }}
  .modal .btns .cancel {{ background: #f2f3f5; color: #4e5969; }}
</style>
</head>
<body>
<div class="wrap">
  <div class="topbar">
    <div><a class="back" href="/">← 返回项目列表</a></div>
    <div class="user">👤 {user} <span class="nav"><a href="/logout">退出</a></span></div>
  </div>

  <div class="card">
    <div class="pcode">{code}</div>
    <div class="pname">{name} {edit_proj_btn}</div>
    <div class="pmeta">创建时间:{created} · BOM: {bom_status}</div>
    {err}
    {upload_bom_html}
  </div>

  <div class="card">
    <div class="card-head">
      <h2>📋 总BOM清单({bom_total} 项)</h2>
      {add_mat_btn}
    </div>
    <div class="tip">💡 点击物料号进入该物料的资料页</div>
    <table>
      <thead><tr><th style="width:22%">物料号</th><th>名称</th><th>规格型号</th><th>制造商</th><th style="width:8%">版本号</th><th style="width:8%">资料数</th><th style="width:12%">操作</th></tr></thead>
      <tbody>{bom_rows}</tbody>
    </table>
  </div>
</div>

{add_mat_modal}
{edit_proj_modal}
{edit_mat_modal}
<script>
function openModal() {{ document.getElementById('mask').classList.add('show'); }}
function closeModal() {{ document.getElementById('mask').classList.remove('show'); }}
function openProjModal() {{ document.getElementById('pj-mask').classList.add('show'); }}
function openMatModal(id) {{
  // 打开编辑物料弹窗,并预填该行数据(数据存在 data-* 属性里)
  var btn = document.getElementById('mat-' + id);
  document.getElementById('e_item').value = btn.getAttribute('data-item');
  document.getElementById('e_old_item').value = btn.getAttribute('data-item');
  document.getElementById('e_name').value = btn.getAttribute('data-name');
  document.getElementById('e_spec').value = btn.getAttribute('data-spec') || '';
  document.getElementById('e_mfr').value = btn.getAttribute('data-mfr') || '';
  document.getElementById('e_ver').value = btn.getAttribute('data-ver') || 'A00';
  document.getElementById('mat-mask').classList.add('show');
}}
function closeAllModals() {{
  ['mask','pj-mask','mat-mask'].forEach(function(id) {{
    var el = document.getElementById(id);
    if (el) el.classList.remove('show');
  }});
}}
['mask','pj-mask','mat-mask'].forEach(function(id) {{
  var el = document.getElementById(id);
  if (el) el.addEventListener('click', function(e) {{
    if (e.target === this) this.classList.remove('show');
  }});
}});
</script>
</body>
</html>"""

MATERIAL_PAGE = """<!DOCTYPE html>
<html lang="zh">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{item} · 物料资料</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: "Microsoft YaHei", "PingFang SC", sans-serif;
         background: #f0f2f5; color: #1f2329; padding: 24px 16px; }}
  .wrap {{ max-width: 900px; margin: 0 auto; }}
  .topbar {{ display: flex; justify-content: space-between; align-items: center; margin-bottom: 14px; }}
  .topbar .back {{ font-size: 13px; color: #3370ff; text-decoration: none; }}
  .topbar .user {{ font-size: 13px; color: #4e5969; }}
  .card {{ background: #fff; border-radius: 10px; box-shadow: 0 2px 12px rgba(0,0,0,.06);
          padding: 24px 28px; margin-bottom: 16px; }}
  .card h2 {{ font-size: 16px; margin-bottom: 14px; }}
  .icode {{ font-size: 15px; color: #3370ff; font-family: Consolas,monospace; font-weight: bold; }}
  .iname {{ font-size: 15px; margin-top: 4px; }}
  .pname {{ font-size: 12px; color: #86909c; margin-top: 4px; }}
  .tip {{ background: #f0f7ff; color: #2f6fed; font-size: 12px; padding: 8px 12px;
         border-radius: 6px; margin-bottom: 12px; }}
  .msg {{ font-size: 13px; padding: 8px 12px; border-radius: 6px; margin-bottom: 12px; }}
  .msg-ok {{ background: #e8f7ef; color: #23a65a; }}
  .msg-err {{ background: #fff3f0; color: #d54941; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 14px; }}
  th {{ text-align: left; color: #86909c; font-weight: 500; font-size: 12px;
       padding: 8px 10px; border-bottom: 1px solid #e5e6eb; }}
  td {{ padding: 9px 10px; border-bottom: 1px solid #f2f3f5; }}
  tr:hover td {{ background: #f7f8fa; }}
  .dl {{ color: #3370ff; text-decoration: none; }}
  .empty {{ color: #86909c; text-align: center; padding: 24px 0; }}
  .upload {{ background: #f7f8fa; border: 1px dashed #c9cdd4; border-radius: 8px;
             padding: 16px; margin-top: 14px; }}
  .upload h3 {{ font-size: 14px; margin-bottom: 10px; color: #4e5969; }}
  .upload input[type=file] {{ font-size: 13px; }}
  .btn {{ display: inline-block; padding: 6px 16px; font-size: 13px; border: none;
         border-radius: 5px; cursor: pointer; color: #fff; background: #23a65a; }}
  .footer {{ margin-top: 16px; font-size: 12px; color: #c0c4cc; text-align: center; }}
</style>
</head>
<body>
<div class="wrap">
  <div class="topbar">
    <div><a class="back" href="/project?code={code_enc}">← 返回项目BOM</a></div>
    <div class="user">👤 {user}</div>
  </div>

  <div class="card">
    <div class="icode">{item}</div>
    <div class="iname">{name}</div>
    <div class="pname">项目:{pname}({code})</div>
  </div>

  <div class="card">
    <h2>📄 资料文件({count} 个)</h2>
    <div class="tip">💡 点击文件名下载。有需要的资料请上传共享。</div>
    {msg}
    <table>
      <thead><tr><th>文件</th><th>大小</th><th>修改时间</th><th>操作</th></tr></thead>
      <tbody>{file_rows}</tbody>
    </table>
  </div>

  {upload_card}

  <div class="footer">技术部项目资料数据库 v4</div>
</div>
</body>
</html>"""

ADMIN_PAGE = """<!DOCTYPE html>
<html lang="zh">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>管理后台 · 技术部项目资料数据库</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: "Microsoft YaHei", "PingFang SC", sans-serif;
         background: #f0f2f5; color: #1f2329; padding: 24px 16px; }}
  .wrap {{ max-width: 1080px; margin: 0 auto; }}
  .topbar {{ display: flex; justify-content: space-between; align-items: center; margin-bottom: 14px; }}
  .topbar .user {{ font-size: 13px; color: #4e5969; }}
  .nav a {{ font-size: 13px; color: #3370ff; text-decoration: none; margin-left: 14px; }}
  .card {{ background: #fff; border-radius: 10px; box-shadow: 0 2px 12px rgba(0,0,0,.06);
          padding: 24px 28px; margin-bottom: 16px; }}
  .card h2 {{ font-size: 16px; margin-bottom: 14px; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
  th {{ text-align: left; color: #86909c; font-weight: 500; font-size: 12px;
       padding: 8px 10px; border-bottom: 1px solid #e5e6eb; }}
  td {{ padding: 8px 10px; border-bottom: 1px solid #f2f3f5; }}
  .empty {{ color: #86909c; text-align: center; padding: 20px 0; }}
  form.inline {{ display: inline; }}
  .btn {{ display: inline-block; padding: 5px 14px; font-size: 13px; border: none;
         border-radius: 5px; cursor: pointer; color: #fff; text-decoration: none;
         background: #3370ff; }}
  .btn:hover {{ opacity: .85; }}
  .approve {{ background: #23a65a; }}
  .reject {{ background: #f53f3f; }}
  .log-up {{ color: #23a65a; font-weight: bold; }}
  .log-dl {{ color: #3370ff; font-weight: bold; }}
  .mono {{ font-family: Consolas,monospace; font-size: 12px; }}
</style>
</head>
<body>
<div class="wrap">
  <div class="topbar">
    <div class="user">👤 管理员:{admin} <span class="nav"><a href="/">← 返回</a><a href="/logout">退出</a></span></div>
  </div>

  <div class="card">
    <h2>📋 待审批的注册申请</h2>
    {pending_html}
  </div>

  <div class="card">
    <h2>📊 操作日志(上传/下载记录)</h2>
    <div class="sub" style="font-size:12px;color:#86909c;margin-bottom:10px">最近 {log_total} 条记录,按时间倒序</div>
    <table>
      <thead><tr><th>时间</th><th>用户</th><th>操作</th><th>项目</th><th>物料</th><th>文件</th><th>大小</th></tr></thead>
      <tbody>{log_rows}</tbody>
    </table>
  </div>

  <div class="card">
    <h2>👥 账号管理</h2>
    <div class="sub" style="font-size:12px;color:#86909c;margin-bottom:10px">共 {acct_count} 个账号 · 仅 admin 可管理 · 其他账号为二级只读</div>
    <table>
      <thead><tr><th>账号</th><th>姓名</th><th>部门</th><th>角色</th><th>密码</th><th>操作</th></tr></thead>
      <tbody>{acct_rows}</tbody>
    </table>
  </div>
</div>
</body>
</html>"""


# ==================== HTTP 处理 ====================

class Handler(BaseHTTPRequestHandler):
    server_version = "ShareServer/4.0"

    def _send(self, body: bytes, ctype: str = "text/html; charset=utf-8", code: int = 200, headers: dict = None):
        self.send_response(code)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(len(body)))
        if headers:
            for k, v in headers.items():
                self.send_header(k, v)
        self.end_headers()
        self.wfile.write(body)

    def _redirect(self, location: str, headers: dict = None):
        self.send_response(302)
        self.send_header("Location", location)
        if headers:
            for k, v in headers.items():
                self.send_header(k, v)
        self.send_header("Content-Length", "0")
        self.end_headers()

    def _html_escape(self, s: str) -> str:
        return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")

    def _safe_join(self, base: str, rel: str) -> str | None:
        """把相对路径拼到 base 下,防路径穿越"""
        rel = unquote(rel)
        full = os.path.normpath(os.path.join(base, rel))
        if not full.startswith(os.path.normpath(base) + os.sep) and full != os.path.normpath(base):
            return None
        return full

    def _find_item_dirs(self, mat_dir: str, item: str) -> list:
        """找到物料号对应的所有资料文件夹(可能同时存在多个)。
        返回列表,例如:
          - [mat_dir/8S131100AA]  (只有纯物料号文件夹)
          - [mat_dir/8S131100AA-总进线智能电表]  (只有 物料号-名称 文件夹)
          - [mat_dir/8S131100AA, mat_dir/8S131100AA-总进线智能电表] (两者都有)
        没有则返回 []。
        """
        dirs = []
        exact = os.path.join(mat_dir, item)
        if os.path.isdir(exact):
            dirs.append(exact)
        # 找 "物料号-" 或 "物料号_" 开头的文件夹
        try:
            prefix1, prefix2 = item + "-", item + "_"
            for fn in sorted(os.listdir(mat_dir)):
                if fn.startswith(prefix1) or fn.startswith(prefix2):
                    d = os.path.join(mat_dir, fn)
                    if d not in dirs and os.path.isdir(d):
                        dirs.append(d)
        except OSError:
            pass
        return dirs

    def _item_upload_dir(self, mat_dir: str, item: str) -> str:
        """上传时的目标文件夹:优先"物料号-名称"文件夹(同事的命名习惯),
        没有则用纯物料号文件夹。
        """
        dirs = self._find_item_dirs(mat_dir, item)
        if not dirs:
            return os.path.join(mat_dir, item)  # 新建纯物料号文件夹
        # 优先带名称的(不含纯物料号),其次第一个
        for d in dirs:
            if os.path.basename(d) != item:
                return d
        return dirs[0]

    def _serve_static(self, path: str):
        STATIC_DIR = os.path.join(BASE_DIR, "static")
        name = os.path.basename(path)
        full = os.path.normpath(os.path.join(STATIC_DIR, name))
        if not full.startswith(os.path.normpath(STATIC_DIR) + os.sep):
            self.send_error(403, "Forbidden")
            return
        if not os.path.isfile(full):
            self.send_error(404, "Not Found")
            return
        ext = os.path.splitext(full)[1].lower()
        ctype = {".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
                 ".gif": "image/gif", ".svg": "image/svg+xml", ".ico": "image/x-icon",
                 ".css": "text/css; charset=utf-8", ".js": "application/javascript",
                 }.get(ext, "application/octet-stream")
        with open(full, "rb") as f:
            data = f.read()
        self._send(data, ctype=ctype)

    # ---------- 登录会话 ----------
    def _get_cookie(self, name: str) -> str | None:
        raw = self.headers.get("Cookie", "")
        for part in raw.split(";"):
            part = part.strip()
            if part.startswith(name + "="):
                return part[len(name) + 1:]
        return None

    def _current_user(self) -> str | None:
        token = self._get_cookie("share_sid")
        if not token:
            return None
        now = time.time()
        with LOCK:
            sess = SESSIONS.get(token)
            if sess is None:
                return None
            if sess["expire_at"] < now:
                SESSIONS.pop(token, None)
                return None
        return sess["user"]

    def _is_admin(self, user: str) -> bool:
        """管理员判定:只有账号名是 admin 才能管理后台、编辑项目/物料。
        硬性绑定账号名(不依赖 role 字段),防止被误升的账号获得管理权。
        """
        with LOCK:
            data = load_auth()
            acct = data["accounts"].get(user)
            return bool(acct and user == "admin")

    def _can_upload_download(self, user: str) -> bool:
        """上传/下载权限:管理员 或 一级用户 可以;二级用户只读。
        管理员:账号名 admin;一级:role == lv1;二级:role == lv2。
        """
        if self._is_admin(user):
            return True
        with LOCK:
            data = load_auth()
            acct = data["accounts"].get(user)
            return bool(acct and acct.get("role") == "lv1")

    def _user_level(self, user: str) -> str:
        """返回用户级别:admin / lv1 / lv2"""
        if self._is_admin(user):
            return "admin"
        with LOCK:
            data = load_auth()
            acct = data["accounts"].get(user)
            return (acct or {}).get("role", "lv2")

    def _user_info(self, user: str) -> dict:
        data = load_auth()
        return data["accounts"].get(user, {})

    # ---------- 路由 ----------
    def do_GET(self):
        path = self.path.split("?", 1)[0]
        user = self._current_user()
        if path.startswith("/static/"):
            self._serve_static(path)
            return
        if path == "/login":
            if user:
                self._redirect("/")
            else:
                self._login_page(msg="")
        elif path == "/logout":
            token = self._get_cookie("share_sid")
            if token:
                with LOCK:
                    SESSIONS.pop(token, None)
            self._redirect("/login")
        elif path == "/register":
            self._register_page(msg="")
        elif path == "/admin":
            if user is None:
                self._redirect("/login")
            elif self._is_admin(user):
                self._admin_page(user)
            else:
                self._send("<html><body><h3>403 无权限</h3><p><a href='/'>返回</a></p></body></html>".encode("utf-8"), code=403)
        elif user is None:
            self._redirect("/login")
        elif path == "/" or path == "/index.html":
            self._index(user)
        elif path == "/project":
            self._project_page(user)
        elif path == "/material":
            self._material_page(user)
        elif path == "/download":
            self._download(user)
        else:
            self.send_error(404, "Not Found")

    def do_POST(self):
        path = self.path.split("?", 1)[0]
        user = self._current_user()
        if path == "/login":
            self._do_login()
        elif path == "/register":
            self._do_register()
        elif path in ("/approve", "/reject"):
            if user is None:
                self._redirect("/login")
            elif self._is_admin(user):
                self._do_decide(approved=(path == "/approve"))
            else:
                self._send(b"403 Forbidden", code=403)
        elif path in ("/setpass", "/setrole", "/setlevel", "/deluser"):
            if user is None:
                self._redirect("/login")
            elif self._is_admin(user):
                self._do_account_admin(path)
            else:
                self._send(b"403 Forbidden", code=403)
        elif path == "/newproject":
            if user is None:
                self._redirect("/login")
            elif self._can_upload_download(user):  # 管理员或一级用户可新建项目
                self._do_newproject()
            else:
                self._send(b"403 Forbidden", code=403)
        elif path == "/addmaterial":
            if user is None:
                self._redirect("/login")
            elif self._can_upload_download(user):  # 管理员或一级用户可添加物料
                self._do_addmaterial()
            else:
                self._send(b"403 Forbidden", code=403)
        elif path == "/editproject":
            if user is None:
                self._redirect("/login")
            elif self._is_admin(user):
                self._do_editproject()
            else:
                self._send(b"403 Forbidden", code=403)
        elif path == "/editmaterial":
            if user is None:
                self._redirect("/login")
            elif self._is_admin(user):
                self._do_editmaterial()
            else:
                self._send(b"403 Forbidden", code=403)
        elif path == "/uploadbom":
            if user is None:
                self._redirect("/login")
            elif self._can_upload_download(user):
                self._do_upload_bom()
            else:
                self._send(b"403 Forbidden", code=403)
        elif path == "/upload":
            if user is None:
                self._redirect("/login")
            elif self._can_upload_download(user):
                self._do_upload(user)
            else:
                self._send(b"403 Forbidden", code=403)
        else:
            self.send_error(404, "Not Found")

    def _read_form(self) -> dict:
        length = int(self.headers.get("Content-Length") or 0)
        body = self.rfile.read(length).decode("utf-8", "replace")
        return parse_qs(body)

    # ---------- 登录/注册/账号 ----------
    def _do_login(self):
        qs = self._read_form()
        user = (qs.get("user") or [""])[0].strip()
        passwd = (qs.get("pass") or [""])[0]
        data = load_auth()
        acct = data["accounts"].get(user)
        if acct and acct["password"] == passwd:
            token = secrets.token_hex(16)
            with LOCK:
                SESSIONS[token] = {"user": user, "expire_at": time.time() + SESSION_TTL}
            self._redirect("/", headers={
                "Set-Cookie": f"share_sid={token}; Path=/; HttpOnly; Max-Age={SESSION_TTL}"
            })
        else:
            self._login_page(msg='<div class="err">账号或密码错误,请重试</div>')

    def _do_register(self):
        qs = self._read_form()
        user = (qs.get("user") or [""])[0].strip()
        passwd = (qs.get("pass") or [""])[0]
        pass2 = (qs.get("pass2") or [""])[0]
        name = (qs.get("name") or [""])[0].strip()
        dept = (qs.get("dept") or [""])[0].strip()
        reason = (qs.get("reason") or [""])[0].strip()
        err = ""
        if not user or not passwd or not name or not dept:
            err = "账号、密码、姓名、部门为必填项"
        elif len(user) < 2 or len(user) > 20:
            err = "账号长度需在 2~20 个字符之间"
        elif passwd != pass2:
            err = "两次输入的密码不一致"
        elif len(passwd) < 6:
            err = "密码至少 6 位"
        if not err:
            with LOCK:
                data = load_auth()
                if user in data["accounts"]:
                    err = f"账号 {user} 已存在,请直接登录"
                elif any(a["user"] == user for a in data["applications"]):
                    err = f"账号 {user} 已提交过申请,请等待审批"
        if err:
            self._register_page(msg=f'<div class="err">{self._html_escape(err)}</div>')
            return
        with LOCK:
            data = load_auth()
            data["applications"].append({
                "id": secrets.token_hex(8), "user": user, "password": passwd,
                "name": name, "dept": dept, "reason": reason,
                "status": "pending", "created": time.strftime("%Y-%m-%d %H:%M"),
            })
            save_auth(data)
        self._register_page(msg=(
            f'<div class="ok">✅ 申请已提交!账号 <b>{self._html_escape(user)}</b> 等待管理员审批,'
            f'通过后即可 <a href="/login">登录</a>。</div>'
        ))

    def _do_decide(self, approved: bool):
        qs = self._read_form()
        aid = (qs.get("id") or [""])[0]
        if not aid:
            self._redirect("/admin")
            return
        with LOCK:
            data = load_auth()
            for a in data["applications"]:
                if a["id"] == aid and a["status"] == "pending":
                    if approved:
                        data["accounts"][a["user"]] = {
                            "password": a["password"], "name": a["name"],
                            "dept": a["dept"], "role": "lv2",
                            "created": time.strftime("%Y-%m-%d %H:%M"),
                        }
                        a["status"] = "approved"
                    else:
                        a["status"] = "rejected"
                    a["decided_at"] = time.strftime("%Y-%m-%d %H:%M")
                    break
            save_auth(data)
        self._redirect("/admin")

    def _do_account_admin(self, path: str):
        qs = self._read_form()
        target = (qs.get("user") or [""])[0].strip()
        with LOCK:
            data = load_auth()
            acct = data["accounts"].get(target)
            if not acct:
                save_auth(data)
                self._redirect("/admin")
                return
            if path == "/setpass":
                newpass = (qs.get("newpass") or [""])[0]
                if len(newpass) < 6:
                    self._send("<html><body><h3>密码至少 6 位</h3><p><a href='/admin'>返回</a></p></body></html>".encode("utf-8"), code=400)
                    return
                acct["password"] = newpass
            elif path == "/setrole":
                if target == "admin":
                    save_auth(data)
                    self._redirect("/admin")
                    return
                admins = [n for n, a in data["accounts"].items() if a.get("role") == "admin"]
                if acct.get("role") == "admin" and len(admins) <= 1:
                    self._send("<html><body><h3>系统必须至少保留一个管理员</h3><p><a href='/admin'>返回</a></p></body></html>".encode("utf-8"), code=400)
                    return
                acct["role"] = "user" if acct.get("role") == "admin" else "admin"
            elif path == "/setlevel":
                # 管理员设置账号为一级(lv1)或二级(lv2);admin 账号不可改
                if target == "admin":
                    save_auth(data)
                    self._redirect("/admin")
                    return
                level = (qs.get("level") or [""])[0].strip()
                if level not in ("lv1", "lv2"):
                    save_auth(data)
                    self._redirect("/admin")
                    return
                acct["role"] = level
            elif path == "/deluser":
                if target == "admin":
                    save_auth(data)
                    self._redirect("/admin")
                    return
                data["accounts"].pop(target, None)
                for tok, sess in list(SESSIONS.items()):
                    if sess["user"] == target:
                        SESSIONS.pop(tok, None)
            save_auth(data)
        self._redirect("/admin")

    # ---------- 项目 ----------
    def _do_newproject(self):
        """管理员新建项目:创建项目文件夹 + 写入索引"""
        qs = self._read_form()
        code = (qs.get("code") or [""])[0].strip()
        name = (qs.get("name") or [""])[0].strip()
        if not code or not name:
            self._send("<html><body><h3>项目编码和名称不能为空</h3><p><a href='/'>返回</a></p></body></html>".encode("utf-8"), code=400)
            return
        # 编码允许:字母、数字、中文、- _ ( ) / 空格
        # (项目编码常是 编码+项目名称 合体,如 HY2026062301-浙江海港...项目)
        if not re.match(r"^[A-Za-z0-9\u4e00-\u9fff\-_()（）/\s]+$", code):
            self._send("<html><body><h3>项目编码含非法字符(仅允许字母、数字、中文、- _ ( ) / 空格)</h3>"
                       "<p><a href='/'>返回</a></p></body></html>".encode("utf-8"), code=400)
            return
        creator = self._current_user() or "admin"
        with LOCK:
            projects = load_projects()
            if code in projects:
                self._send("<html><body><h3>该项目编码已存在</h3><p><a href='/'>返回</a></p></body></html>".encode("utf-8"), code=400)
                return
            proj_dir = os.path.join(ROOT, code)
            os.makedirs(os.path.join(proj_dir, "物料资料"), exist_ok=True)
            projects[code] = {
                "name": name,
                "created": time.strftime("%Y-%m-%d %H:%M"),
                "creator": creator,
                "bom": None,
            }
            save_projects(projects)
        self._redirect("/")

    def _do_addmaterial(self):
        """管理员在项目页添加物料:写入 BOM Excel(没有则自动创建)+ 建物料文件夹"""
        qs = self._read_form()
        code = (qs.get("project") or [""])[0].strip()
        item = (qs.get("item") or [""])[0].strip()
        name = (qs.get("name") or [""])[0].strip()
        spec = (qs.get("spec") or [""])[0].strip()
        mfr = (qs.get("manufacturer") or [""])[0].strip()

        if not code or not item or not name:
            self._send("<html><body><h3>物料号和名称为必填项</h3>"
                       "<p><a href='/'>返回</a></p></body></html>".encode("utf-8"), code=400)
            return
        if not re.match(r"^[A-Za-z0-9\-_/]+$", item):
            self._send("<html><body><h3>物料号只能含字母、数字、- _ /</h3>"
                       "<p><a href='/'>返回</a></p></body></html>".encode("utf-8"), code=400)
            return

        proj_dir = os.path.join(ROOT, code)
        if not os.path.isdir(proj_dir):
            self._redirect("/")
            return
        mat_dir = os.path.join(proj_dir, "物料资料")
        os.makedirs(mat_dir, exist_ok=True)

        try:
            from openpyxl import load_workbook, Workbook
            bom_path = os.path.join(proj_dir, "总BOM清单.xlsx")
            if os.path.exists(bom_path):
                wb = load_workbook(bom_path)
                ws = wb.active
                # 检查物料号是否已存在
                exists = False
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and str(row[0]).strip() == item:
                        exists = True
                        break
                if exists:
                    wb.close()
                    self._send("<html><body><h3>该物料号已在 BOM 中</h3>"
                               "<p><a href='/'>返回</a></p></body></html>".encode("utf-8"), code=400)
                    return
                # 确保表头有新列(旧 BOM 自动补 制造商/版本号),按列写入
                cols = ensure_bom_columns(ws)
                next_row = ws.max_row + 1
                ws.cell(row=next_row, column=cols["code"]).value = item
                ws.cell(row=next_row, column=cols["name"]).value = name
                if cols["spec"]:
                    ws.cell(row=next_row, column=cols["spec"]).value = spec
                if cols["mfr"]:
                    ws.cell(row=next_row, column=cols["mfr"]).value = mfr
                if cols["ver"]:
                    ws.cell(row=next_row, column=cols["ver"]).value = "A00"
            else:
                # 没有 BOM 文件,自动创建一个带表头的
                wb = Workbook()
                ws = wb.active
                ws.title = "总BOM清单"
                ws.append(["物料号", "名称", "规格型号", "制造商", "版本号", "数量", "单位"])
                ws.append([item, name, spec, mfr, "A00", "", ""])
            wb.save(bom_path)
            wb.close()
            # 建物料文件夹(物料号-名称,符合同事命名习惯)
            os.makedirs(os.path.join(mat_dir, f"{item}-{name}"), exist_ok=True)
            # 更新项目索引的 bom 标记
            with LOCK:
                projects = load_projects()
                if code in projects:
                    projects[code]["bom"] = "总BOM清单.xlsx"
                    save_projects(projects)
        except Exception as e:
            err_msg = str(e)
            if "Permission denied" in err_msg or "拒绝访问" in err_msg:
                hint = ("BOM 文件被占用(可能正用 WPS/Excel 打开着 总BOM清单.xlsx)。"
                        "请先关闭该文件再重试。")
            else:
                hint = err_msg
            self._send(f"<html><body><h3>添加物料失败</h3><p>{self._html_escape(hint)}</p>"
                       f"<p><a href='javascript:history.back()'>返回</a></p></body></html>".encode("utf-8"), code=500)
            return
        self._redirect(f"/project?code={quote(code)}")

    def _do_editproject(self):
        """管理员修改项目名称"""
        qs = self._read_form()
        code = (qs.get("code") or [""])[0].strip()
        name = (qs.get("name") or [""])[0].strip()
        if not code or not name:
            self._send("<html><body><h3>项目名称不能为空</h3>"
                       "<p><a href='/'>返回</a></p></body></html>".encode("utf-8"), code=400)
            return
        with LOCK:
            projects = load_projects()
            if code not in projects:
                self._send("<html><body><h3>项目不存在</h3>"
                           "<p><a href='/'>返回</a></p></body></html>".encode("utf-8"), code=404)
                return
            projects[code]["name"] = name
            save_projects(projects)
        self._redirect(f"/project?code={quote(code)}")

    def _do_editmaterial(self):
        """管理员修改物料信息:物料号/名称/规格/制造商/版本号
        物料号或名称变更时,同步重命名物料资料文件夹,避免资料"丢失"。
        """
        qs = self._read_form()
        code = (qs.get("project") or [""])[0].strip()
        old_item = (qs.get("old_item") or [""])[0].strip()
        new_item = (qs.get("item") or [""])[0].strip()
        name = (qs.get("name") or [""])[0].strip()
        spec = (qs.get("spec") or [""])[0].strip()
        mfr = (qs.get("manufacturer") or [""])[0].strip()
        ver = (qs.get("version") or [""])[0].strip()

        if not code or not old_item or not new_item or not name:
            self._send("<html><body><h3>物料号和名称为必填项</h3>"
                       "<p><a href='/'>返回</a></p></body></html>".encode("utf-8"), code=400)
            return
        if not re.match(r"^[A-Za-z0-9\-_/]+$", new_item):
            self._send("<html><body><h3>物料号只能含字母、数字、- _ /</h3>"
                       "<p><a href='/'>返回</a></p></body></html>".encode("utf-8"), code=400)
            return

        proj_dir = os.path.join(ROOT, code)
        if not os.path.isdir(proj_dir):
            self._redirect("/")
            return
        bom_path = os.path.join(proj_dir, "总BOM清单.xlsx")
        if not os.path.exists(bom_path):
            self._send("<html><body><h3>该项目还没有 BOM</h3>"
                       "<p><a href='/'>返回</a></p></body></html>".encode("utf-8"), code=400)
            return

        # 1. 更新 BOM Excel 里的该物料行
        try:
            from openpyxl import load_workbook
            wb = load_workbook(bom_path)
            ws = wb.active
            cols = ensure_bom_columns(ws)
            code_col, name_col = cols["code"], cols["name"]
            found = False
            for r in range(2, ws.max_row + 1):
                if str(ws.cell(row=r, column=code_col).value or "").strip() == old_item:
                    ws.cell(row=r, column=code_col).value = new_item
                    ws.cell(row=r, column=name_col).value = name
                    if cols["spec"]:
                        ws.cell(row=r, column=cols["spec"]).value = spec
                    if cols["mfr"]:
                        ws.cell(row=r, column=cols["mfr"]).value = mfr
                    if cols["ver"] and ver:
                        ws.cell(row=r, column=cols["ver"]).value = ver
                    found = True
                    break
            if not found:
                wb.close()
                self._send("<html><body><h3>物料号不存在于 BOM</h3>"
                           "<p><a href='/'>返回</a></p></body></html>".encode("utf-8"), code=404)
                return
            wb.save(bom_path)
            wb.close()
        except Exception as e:
            err_msg = str(e)
            if "Permission denied" in err_msg or "拒绝访问" in err_msg:
                hint = ("BOM 文件被占用(可能正用 WPS/Excel 打开着 总BOM清单.xlsx)。"
                        "请先关闭该文件再重试。")
            else:
                hint = err_msg
            self._send(f"<html><body><h3>修改物料失败</h3><p>{self._html_escape(hint)}</p>"
                       f"<p><a href='javascript:history.back()'>返回</a></p></body></html>".encode("utf-8"), code=500)
            return

        # 2. 同步重命名物料资料文件夹(物料号变了或名称变了)
        mat_dir = os.path.join(proj_dir, "物料资料")
        try:
            old_dirs = self._find_item_dirs(mat_dir, old_item)
            new_folder_name = f"{new_item}-{name}" if new_item != name else new_item
            # 目标文件夹:优先用 "新物料号-新名称",若已存在则用纯物料号
            target = os.path.join(mat_dir, new_folder_name)
            if os.path.isdir(target):
                target = os.path.join(mat_dir, new_item)
            for d in old_dirs:
                if os.path.normpath(d) == os.path.normpath(target):
                    continue  # 已是目标,不用动
                if os.path.exists(target):
                    # 目标已存在:把旧文件夹里的文件合并进去
                    for root, dirs, files in os.walk(d):
                        rel = os.path.relpath(root, d)
                        dst_dir = os.path.join(target, rel) if rel != "." else target
                        os.makedirs(dst_dir, exist_ok=True)
                        for fn in files:
                            src_f = os.path.join(root, fn)
                            dst_f = os.path.join(dst_dir, fn)
                            if not os.path.exists(dst_f):
                                os.replace(src_f, dst_f)
                    if not os.listdir(d):
                        os.rmdir(d)
                    elif d != os.path.join(mat_dir, old_item):
                        # 旧文件夹还有内容且不是纯物料号文件夹,保留但改名
                        pass
                else:
                    os.rename(d, target)
        except Exception:
            pass  # 文件夹改名失败不影响 BOM 修改

        self._redirect(f"/project?code={quote(code)}")

    def _do_upload_bom(self):
        """管理员上传项目的 BOM Excel(保存为 总BOM清单.xlsx)"""
        length = int(self.headers.get("Content-Length") or 0)
        if length > MAX_UPLOAD:
            self._send(b"file too large", code=413)
            return
        body = self.rfile.read(length)
        ctype = self.headers.get("Content-Type", "")
        fields, files = parse_multipart(body, ctype)
        code = (fields.get("project") or "").strip()
        fname, fdata = files.get("file", ("", b""))
        if not code or not fdata:
            self._redirect("/project?code=" + quote(code))
            return
        if not fname.lower().endswith((".xlsx", ".xlsm")):
            self._send("<html><body><h3>请上传 .xlsx 格式的 Excel 文件</h3><p><a href='/'>返回</a></p></body></html>".encode("utf-8"), code=400)
            return
        proj_dir = os.path.join(ROOT, code)
        if not os.path.isdir(proj_dir):
            self._redirect("/")
            return
        bom_path = os.path.join(proj_dir, "总BOM清单.xlsx")
        with open(bom_path, "wb") as f:
            f.write(fdata)
        with LOCK:
            projects = load_projects()
            if code in projects:
                projects[code]["bom"] = "总BOM清单.xlsx"
                save_projects(projects)
        self._redirect("/project?code=" + quote(code))

    # ---------- 页面渲染 ----------
    def _login_page(self, msg: str):
        self._send(LOGIN_PAGE.format(msg=msg).encode("utf-8"))

    def _register_page(self, msg: str):
        self._send(REGISTER_PAGE.format(msg=msg).encode("utf-8"))

    def _index(self, user: str):
        """首页:只显示项目列表,双击进入"""
        projects = load_projects()
        rows = []
        for i, (code, p) in enumerate(sorted(projects.items()), start=1):
            # 名称去重:编码是"前缀-项目名"合体(HY2026062301-项目名),
            # 所以名称行只显示项目名,编码行只显示前缀,不重复
            name = p.get("name", "") or ""
            code_part = code.split("-")[0]  # 编码前缀,如 HY2026062301
            # 若 name 为空或就是编码本身,则从编码里剥离前缀当名称
            if not name or name == code:
                name = code[len(code_part):].lstrip("-") or code
            bom_ok = bool(p.get("bom"))
            badge = ('<span class="badge badge-ok">有BOM</span>' if bom_ok
                     else '<span class="badge badge-no">未上传BOM</span>')
            creator = p.get("creator", "—")
            rows.append(
                f'<tr class="proj" ondblclick="location.href=\'/project?code={quote(code)}\'">'
                f'<td class="idx">{i}</td>'
                f'<td><div class="pname">{self._html_escape(name)}</div>'
                f'<div class="pcode">{self._html_escape(code_part)}</div></td>'
                f'<td class="pcreator">👤 {self._html_escape(creator)}</td>'
                f'<td>{badge}</td>'
                f'<td class="pmeta">{self._html_escape(p.get("created",""))}</td>'
                f'<td class="hint">双击进入 ➜</td>'
                f'</tr>'
            )
        if rows:
            projects_html = (
                '<table class="projlist">'
                '<thead><tr><th style="width:5%">#</th><th>项目</th>'
                '<th style="width:14%">创建者</th><th style="width:12%">BOM</th>'
                '<th style="width:16%">创建时间</th><th style="width:10%"></th></tr></thead>'
                f'<tbody>{"".join(rows)}</tbody></table>'
            )
        else:
            projects_html = '<div class="empty">还没有项目,管理员可点击右上角「新建项目」创建</div>'

        # 管理员或一级用户:可新建项目;管理后台入口仅管理员
        if self._is_admin(user):
            with LOCK:
                data = load_auth()
                pend_cnt = sum(1 for a in data["applications"] if a["status"] == "pending")
            admin_nav = f'<a href="/admin">管理后台{("("+str(pend_cnt)+")") if pend_cnt else ""}</a><a href="/logout">退出</a>'
        else:
            admin_nav = '<a href="/logout">退出</a>'

        if self._can_upload_download(user):
            newproj_btn = '<button class="btn btn-green" onclick="openModal()">➕ 新建项目</button>'
            newproj_modal = (
                '<div class="mask" id="mask"><div class="modal">'
                '<h3>➕ 新建项目</h3>'
                '<form method="post" action="/newproject">'
                '<label>项目编码</label>'
                '<input type="text" name="code" placeholder="如 HY2026062301" required>'
                '<label>项目名称</label>'
                '<input type="text" name="name" placeholder="如:XX项目" required>'
                '<div class="btns"><button type="button" class="cancel" onclick="closeModal()">取消</button>'
                '<button type="submit" class="ok">创建项目</button></div>'
                '</form></div></div>'
            )
        else:
            newproj_btn = ""
            newproj_modal = ""

        page = HOME_PAGE.format(
            user=self._html_escape(user),
            admin_nav=admin_nav,
            newproj_btn=newproj_btn,
            newproj_modal=newproj_modal,
            projects_html=projects_html,
        )
        self._send(page.encode("utf-8"))

    def _project_page(self, user: str):
        qs = parse_qs(urlparse(self.path).query)
        code = (qs.get("code") or [""])[0]
        projects = load_projects()
        if code not in projects:
            self._send("<html><body><h3>项目不存在</h3>"
                    "<p><a href='/'>返回</a></p></body></html>".encode("utf-8"), code=404)
            return
        p = projects[code]
        proj_dir = os.path.join(ROOT, code)
        mat_dir = os.path.join(proj_dir, "物料资料")
        is_admin = self._is_admin(user)

        # BOM
        bom_path = os.path.join(proj_dir, "总BOM清单.xlsx")
        bom_items, bom_err = parse_bom_excel(bom_path) if p.get("bom") and os.path.exists(bom_path) else (None, "尚未上传 BOM Excel")

        if bom_items is None:
            # 没有 BOM 或解析失败
            err_html = f'<div class="err">{self._html_escape(bom_err)}</div>' if bom_err else ""
            bom_rows = f'<tr><td colspan="7" class="empty">{self._html_escape(bom_err or "无 BOM")}</td></tr>'
            bom_total = 0
            bom_status = "未上传"
        else:
            err_html = ""
            bom_total = len(bom_items)
            bom_status = f"总BOM清单.xlsx · {bom_total} 项"
            rows = []
            for m in bom_items:
                code_item = m["code"]
                # 该物料所有资料文件夹(可能多个)下的资料数
                dirs = self._find_item_dirs(mat_dir, code_item)
                files = []
                for d in dirs:
                    files.extend(list_files(d))
                cnt = len(files)
                # 管理员可编辑该物料
                edit_btn = ""
                if is_admin:
                    edit_btn = (
                        f'<button class="btn btn-edit" style="padding:3px 10px;font-size:12px;margin-left:6px" '
                        f'id="mat-{quote(code_item)}" '
                        f'data-item="{self._html_escape(code_item)}" '
                        f'data-name="{self._html_escape(m["name"] or "")}" '
                        f'data-spec="{self._html_escape(m.get("spec") or "")}" '
                        f'data-mfr="{self._html_escape(m.get("manufacturer") or "")}" '
                        f'data-ver="{self._html_escape(m.get("version") or "A00")}" '
                        f'onclick="openMatModal(\'{quote(code_item)}\')">✏ 编辑</button>'
                    )
                rows.append(
                    f'<tr><td class="code">{self._html_escape(code_item)}</td>'
                    f'<td class="mat"><a href="/material?code={quote(code)}&item={quote(code_item)}">'
                    f'{self._html_escape(m["name"] or "—")}</a>{edit_btn}</td>'
                    f'<td>{self._html_escape(m.get("spec") or "—")}</td>'
                    f'<td>{self._html_escape(m.get("manufacturer") or "—")}</td>'
                    f'<td><span class="ver">{self._html_escape(m.get("version") or "A00")}</span></td>'
                    f'<td class="cnt">{cnt}</td>'
                    f'<td><a class="btn" style="padding:4px 12px;font-size:12px" '
                    f'href="/material?code={quote(code)}&item={quote(code_item)}">查看资料</a></td></tr>'
                )
            bom_rows = "".join(rows)

        # 管理员或一级用户:可上传BOM、添加物料
        if self._can_upload_download(user):
            upload_bom_html = (
                '<form class="upload" method="post" action="/uploadbom" enctype="multipart/form-data">'
                f'<input type="hidden" name="project" value="{self._html_escape(code)}">'
                '<input type="file" name="file" accept=".xlsx" required>'
                '<button class="btn btn-green" type="submit">上传/更新 BOM Excel</button>'
                '</form>'
            )
            add_mat_btn = '<button class="btn btn-green" onclick="openModal()">➕ 添加物料</button>'
            add_mat_modal = (
                '<div class="mask" id="mask"><div class="modal">'
                '<h3>➕ 添加物料到 BOM</h3>'
                '<form method="post" action="/addmaterial">'
                f'<input type="hidden" name="project" value="{self._html_escape(code)}">'
                '<label>物料号</label>'
                '<input type="text" name="item" placeholder="如 8S131100AA" required>'
                '<label>物料名称</label>'
                '<input type="text" name="name" placeholder="如:总进线智能电表" required>'
                '<label>规格型号(可选)</label>'
                '<input type="text" name="spec" placeholder="如 DTSD341-MC3 三相四线">'
                '<label>制造商(可选)</label>'
                '<input type="text" name="manufacturer" placeholder="如:正泰、德力西">'
                '<div class="btns"><button type="button" class="cancel" onclick="closeModal()">取消</button>'
                '<button type="submit" class="ok">添加物料</button></div>'
                '</form></div></div>'
            )
        else:
            upload_bom_html = ""
            add_mat_btn = ""
            add_mat_modal = ""
            edit_proj_btn = ""
            edit_proj_modal = ""
            edit_mat_modal = ""

        edit_proj_btn = ""
        edit_proj_modal = ""
        edit_mat_modal = ""
        if self._is_admin(user):
            edit_proj_btn = ('<button class="btn btn-edit" style="padding:3px 12px;font-size:12px" '
                             'onclick="openProjModal()">✏ 编辑</button>')
            edit_proj_modal = (
                '<div class="mask" id="pj-mask"><div class="modal">'
                '<h3>✏ 编辑项目名称</h3>'
                '<form method="post" action="/editproject">'
                f'<input type="hidden" name="code" value="{self._html_escape(code)}">'
                '<label>项目名称</label>'
                f'<input type="text" name="name" value="{self._html_escape(p.get("name",""))}" required>'
                '<div class="btns"><button type="button" class="cancel" onclick="closeAllModals()">取消</button>'
                '<button type="submit" class="ok">保存</button></div>'
                '</form></div></div>'
            )
            edit_mat_modal = (
                '<div class="mask" id="mat-mask"><div class="modal">'
                '<h3>✏ 编辑物料信息</h3>'
                '<form method="post" action="/editmaterial">'
                f'<input type="hidden" name="project" value="{self._html_escape(code)}">'
                '<input type="hidden" name="old_item" id="e_old_item">'
                '<label>物料号</label>'
                '<input type="text" name="item" id="e_item" required>'
                '<label>物料名称</label>'
                '<input type="text" name="name" id="e_name" required>'
                '<label>规格型号</label>'
                '<input type="text" name="spec" id="e_spec">'
                '<label>制造商</label>'
                '<input type="text" name="manufacturer" id="e_mfr">'
                '<label>版本号</label>'
                '<input type="text" name="version" id="e_ver" placeholder="如 A00">'
                '<div class="btns"><button type="button" class="cancel" onclick="closeAllModals()">取消</button>'
                '<button type="submit" class="ok">保存修改</button></div>'
                '</form></div></div>'
            )

        page = PROJECT_PAGE.format(
            code=self._html_escape(code),
            name=self._html_escape(p.get("name", "")),
            created=self._html_escape(p.get("created", "")),
            bom_status=self._html_escape(bom_status),
            err=err_html,
            upload_bom_html=upload_bom_html,
            add_mat_btn=add_mat_btn,
            add_mat_modal=add_mat_modal,
            edit_proj_btn=edit_proj_btn,
            edit_proj_modal=edit_proj_modal,
            edit_mat_modal=edit_mat_modal,
            bom_total=bom_total,
            bom_rows=bom_rows,
            user=self._html_escape(user),
        )
        self._send(page.encode("utf-8"))

    def _material_page(self, user: str):
        qs = parse_qs(urlparse(self.path).query)
        code = (qs.get("code") or [""])[0]
        item = (qs.get("item") or [""])[0]
        msg = (qs.get("msg") or [""])[0]
        projects = load_projects()
        if code not in projects:
            self._send("<html><body><h3>项目不存在</h3>"
                    "<p><a href='/'>返回</a></p></body></html>".encode("utf-8"), code=404)
            return
        p = projects[code]
        proj_dir = os.path.join(ROOT, code)
        mat_dir = os.path.join(proj_dir, "物料资料")
        item_dirs = self._find_item_dirs(mat_dir, item)

        # 从 BOM 里找物料名
        name = item
        bom_path = os.path.join(proj_dir, "总BOM清单.xlsx")
        if p.get("bom") and os.path.exists(bom_path):
            items, _ = parse_bom_excel(bom_path)
            if items:
                for m in items:
                    if m["code"] == item:
                        name = m["name"] or item
                        break

        # 合并所有物料文件夹的文件,显示时带文件夹前缀区分
        files = []
        for d in item_dirs:
            dname = os.path.basename(d)
            for rel, full, size, mtime in list_files(d):
                files.append((f"{dname}/{rel}" if dname != item else rel, full, size, mtime))
        # 管理员/一级用户可下载;二级用户只读
        can_download = self._can_upload_download(user)
        if files:
            rows = []
            for disp, full, size, mtime in sorted(files, key=lambda x: x[0].lower()):
                if can_download:
                    dl = f"/download?proj={quote(code)}&item={quote(item)}&f={quote(disp)}"
                    op = f'<a class="dl" href="{dl}">⬇ 下载</a>'
                else:
                    op = '<span style="color:#c0c4cc;font-size:12px">🔒 仅管理员/一级用户可下载</span>'
                rows.append(
                    f'<tr><td>📄 {self._html_escape(disp)}</td>'
                    f'<td>{fmt_size(size)}</td>'
                    f'<td>{time.strftime("%Y-%m-%d %H:%M", time.localtime(mtime))}</td>'
                    f'<td>{op}</td></tr>'
                )
            file_rows = "".join(rows)
        else:
            file_rows = f'<tr><td colspan="4" class="empty">该物料还没有资料{"，等待一级用户上传" if not can_download else ",上传第一份吧 📤"}</td></tr>'

        msg_html = ""
        if msg == "ok":
            msg_html = '<div class="msg msg-ok">✅ 上传成功!</div>'
        elif msg == "dup":
            msg_html = '<div class="msg msg-err">⚠️ 同名文件已存在,已自动加序号保存</div>'
        elif msg == "err":
            msg_html = '<div class="msg msg-err">❌ 上传失败,请重试</div>'

        # 上传卡片:管理员/一级用户可见
        if self._can_upload_download(user):
            upload_card = (
                '<div class="card"><h2>⬆ 上传资料</h2>'
                '<form class="upload" method="post" action="/upload" enctype="multipart/form-data">'
                f'<input type="hidden" name="project" value="{self._html_escape(code)}">'
                f'<input type="hidden" name="item" value="{self._html_escape(item)}">'
                '<input type="file" name="file" required>'
                '<button class="btn" type="submit">上传到该物料</button>'
                '</form></div>'
            )
        else:
            upload_card = (
                '<div class="card">'
                '<div style="color:#86909c;font-size:13px">🔒 上传权限仅限管理员和一级用户。'
                '如需上传资料,请联系管理员。</div>'
                '</div>'
            )

        page = MATERIAL_PAGE.format(
            code=self._html_escape(code),
            code_enc=quote(code),
            item=self._html_escape(item),
            name=self._html_escape(name),
            pname=self._html_escape(p.get("name", "")),
            user=self._html_escape(user),
            count=len(files),
            msg=msg_html,
            file_rows=file_rows,
            upload_card=upload_card,
        )
        self._send(page.encode("utf-8"))

    # ---------- 下载 / 上传 ----------
    def _download(self, user: str):
        # 二级用户不可下载(管理员/一级用户可以)
        if not self._can_upload_download(user):
            self._send("<html><body><h3>403 无权限:下载仅限管理员和一级用户</h3>"
                       "<p><a href='javascript:history.back()'>返回</a></p></body></html>".encode("utf-8"), code=403)
            return
        qs = parse_qs(urlparse(self.path).query)
        if "proj" in qs:
            # 物料资料下载:proj + item + f
            code = (qs.get("proj") or [""])[0]
            item = (qs.get("item") or [""])[0]
            rel = (qs.get("f") or [""])[0]
            mat_dir = os.path.join(ROOT, code, "物料资料")
            item_dirs = self._find_item_dirs(mat_dir, item)
            # 在多个物料文件夹中找文件。
            # rel 可能是 "图纸/x.pdf"(相对物料文件夹)或 "8S131100AA-名称/图纸/x.pdf"(带文件夹前缀)
            full = None
            for d in item_dirs:
                dname = os.path.basename(d)
                r = rel
                if dname != item and r.startswith(dname + "/"):
                    r = r[len(dname) + 1:]  # 剥离文件夹前缀
                cand = self._safe_join(d, r)
                if cand and os.path.isfile(cand):
                    full = cand
                    break
            if full is None:
                self.send_error(404, "File not found")
                return
            log_activity(user, "下载", code, item, os.path.basename(full), os.path.getsize(full))
        else:
            # 未归档文件下载(仅根目录直属文件)
            fn = (qs.get("f") or [""])[0]
            full = self._safe_join(ROOT, fn)
            if full is None or not os.path.isfile(full) or os.path.dirname(full) != os.path.normpath(ROOT):
                self.send_error(404, "File not found")
                return
            log_activity(user, "下载", "(未归档)", "", os.path.basename(full), os.path.getsize(full))

        size = os.path.getsize(full)
        fname = os.path.basename(full)
        ascii_name = fname.encode("ascii", "ignore").decode() or "download"
        disp = f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{quote(fname)}"
        self.send_response(200)
        self.send_header("Content-Type", "application/octet-stream")
        self.send_header("Content-Disposition", disp)
        self.send_header("Content-Length", str(size))
        self.end_headers()
        with open(full, "rb") as fh:
            while True:
                chunk = fh.read(65536)
                if not chunk:
                    break
                self.wfile.write(chunk)

    def _do_upload(self, user: str):
        """上传资料到物料文件夹"""
        length = int(self.headers.get("Content-Length") or 0)
        if length > MAX_UPLOAD:
            self._send(b"file too large", code=413)
            return
        body = self.rfile.read(length)
        ctype = self.headers.get("Content-Type", "")
        fields, files = parse_multipart(body, ctype)
        code = (fields.get("project") or "").strip()
        item = (fields.get("item") or "").strip()
        fname, fdata = files.get("file", ("", b""))

        if not code or not item or not fdata:
            self._redirect(f"/material?code={quote(code)}&item={quote(item)}&msg=err")
            return

        mat_dir = os.path.join(ROOT, code, "物料资料")
        item_dir = self._item_upload_dir(mat_dir, item)
        try:
            os.makedirs(item_dir, exist_ok=True)
            # 同名文件自动加序号,避免覆盖
            fname = os.path.basename(fname)  # 防路径穿越
            fname = re.sub(r'[\\/:*?"<>|]', "_", fname)
            save_path = os.path.join(item_dir, fname)
            dup = False
            if os.path.exists(save_path):
                dup = True
                base, ext = os.path.splitext(fname)
                i = 1
                while os.path.exists(os.path.join(item_dir, f"{base}({i}){ext}")):
                    i += 1
                save_path = os.path.join(item_dir, f"{base}({i}){ext}")
            with open(save_path, "wb") as f:
                f.write(fdata)
            log_activity(user, "上传", code, item, os.path.basename(save_path), len(fdata))
            # 资料有改动 → 该物料版本号 +1(A00 -> A01 -> A02 ...)
            self._bump_material_version(code, item)
            msg = "dup" if dup else "ok"
        except Exception:
            msg = "err"
        self._redirect(f"/material?code={quote(code)}&item={quote(item)}&msg={msg}")

    def _bump_material_version(self, code: str, item: str):
        """资料上传后,把该物料在 BOM 里的版本号 +1(A00->A01->A02...)"""
        try:
            from openpyxl import load_workbook
            proj_dir = os.path.join(ROOT, code)
            bom_path = os.path.join(proj_dir, "总BOM清单.xlsx")
            if not os.path.exists(bom_path):
                return
            wb = load_workbook(bom_path)
            ws = wb.active
            cols = ensure_bom_columns(ws)
            ver_col = cols["ver"]
            code_col = cols["code"]
            # 找物料行
            for r in range(2, ws.max_row + 1):
                if str(ws.cell(row=r, column=code_col).value or "").strip() == item:
                    cur = str(ws.cell(row=r, column=ver_col).value or "A00").strip()
                    ws.cell(row=r, column=ver_col).value = bump_version(cur)
                    break
            wb.save(bom_path)
            wb.close()
        except Exception:
            pass  # 版本号更新失败不影响上传本身

    # ---------- 管理后台 ----------
    def _admin_page(self, admin: str):
        data = load_auth()
        # 待审批
        pendings = [a for a in data["applications"] if a["status"] == "pending"]
        if pendings:
            rows = []
            for a in pendings:
                reason_s = f'<div style="color:#86909c;font-size:12px">理由:{self._html_escape(a.get("reason") or "无")}</div>' if a.get("reason") else ""
                rows.append(
                    f'<tr><td><b>{self._html_escape(a["user"])}</b></td>'
                    f'<td>{self._html_escape(a["name"])}</td>'
                    f'<td>{self._html_escape(a["dept"])}</td>'
                    f'<td>{a["created"]}{reason_s}</td>'
                    f'<td style="white-space:nowrap">'
                    f'<form class="inline" method="post" action="/approve">'
                    f'<input type="hidden" name="id" value="{a["id"]}">'
                    f'<button class="btn approve" type="submit">✓ 同意</button></form> '
                    f'<form class="inline" method="post" action="/reject">'
                    f'<input type="hidden" name="id" value="{a["id"]}">'
                    f'<button class="btn reject" type="submit">✕ 拒绝</button></form>'
                    f'</td></tr>'
                )
            pending_html = (
                '<table><thead><tr><th>账号</th><th>姓名</th><th>部门</th>'
                '<th>申请时间</th><th>操作</th></tr></thead>'
                f'<tbody>{"".join(rows)}</tbody></table>'
            )
        else:
            pending_html = '<div class="empty">🎉 暂无待审批申请</div>'

        # 操作日志
        try:
            with open(ACTIVITY_FILE, "r", encoding="utf-8") as f:
                logs = json.load(f)
        except Exception:
            logs = []
        logs_rev = list(reversed(logs))[:100]
        if logs_rev:
            log_rows = []
            for lg in logs_rev:
                act = lg.get("action", "")
                cls = "log-up" if act == "上传" else "log-dl"
                log_rows.append(
                    f'<tr><td class="mono">{lg.get("time","")}</td>'
                    f'<td>{self._html_escape(lg.get("user",""))}</td>'
                    f'<td class="{cls}">{act}</td>'
                    f'<td class="mono">{self._html_escape(lg.get("project",""))}</td>'
                    f'<td class="mono">{self._html_escape(lg.get("item",""))}</td>'
                    f'<td>{self._html_escape(lg.get("filename",""))}</td>'
                    f'<td>{fmt_size(lg.get("size",0))}</td></tr>'
                )
            log_rows_html = "".join(log_rows)
        else:
            log_rows_html = '<tr><td colspan="7" class="empty">暂无操作记录</td></tr>'

        # 账号管理
        accts = sorted(data["accounts"].items())
        acct_rows = []
        for name, a in accts:
            lv = a.get("role", "lv2")
            role = {"admin": "管理员", "lv1": "一级用户", "lv2": "二级用户"}.get(lv, "二级用户")
            if name == "admin":
                ops = (f'<form class="inline" method="post" action="/setpass">'
                       f'<input type="hidden" name="user" value="{self._html_escape(name)}">'
                       f'<input type="password" name="newpass" placeholder="新密码(≥6位)" size="8">'
                       f'<button class="btn approve" type="submit">改密码</button></form>')
            else:
                # 管理员可把账号设为一级/二级(决定能否上传下载)
                cur = "lv1" if lv == "lv1" else "lv2"
                other = "lv2" if cur == "lv1" else "lv1"
                other_label = "二级" if other == "lv2" else "一级"
                ops = (f'<form class="inline" method="post" action="/setpass">'
                       f'<input type="hidden" name="user" value="{self._html_escape(name)}">'
                       f'<input type="password" name="newpass" placeholder="新密码(≥6位)" size="8">'
                       f'<button class="btn approve" type="submit">改密码</button></form> '
                       f'<form class="inline" method="post" action="/setlevel">'
                       f'<input type="hidden" name="user" value="{self._html_escape(name)}">'
                       f'<input type="hidden" name="level" value="{other}">'
                       f'<button class="btn" type="submit">设为{other_label}用户</button></form> '
                       f'<form class="inline" method="post" action="/deluser" '
                       f'onsubmit="return confirm(\'确定删除账号 {self._html_escape(name)}?\')">'
                       f'<input type="hidden" name="user" value="{self._html_escape(name)}">'
                       f'<button class="btn reject" type="submit">删除</button></form>')
            acct_rows.append(
                f'<tr><td><b>{self._html_escape(name)}</b></td>'
                f'<td>{self._html_escape(a.get("name",""))}</td>'
                f'<td>{self._html_escape(a.get("dept",""))}</td>'
                f'<td>{role}</td>'
                f'<td style="font-family:Consolas,monospace">{self._html_escape(a.get("password",""))}</td>'
                f'<td style="white-space:nowrap">{ops}</td></tr>'
            )
        page = ADMIN_PAGE.format(
            admin=self._html_escape(admin),
            pending_html=pending_html,
            log_total=len(logs_rev),
            log_rows=log_rows_html,
            acct_count=len(accts),
            acct_rows="".join(acct_rows) or '<tr><td colspan="6" class="empty">暂无账号</td></tr>',
        )
        self._send(page.encode("utf-8"))

    def log_message(self, fmt, *args):
        try:
            sys.stdout.write("[%s] %s\n" % (self.log_date_time_string(), fmt % args))
            sys.stdout.flush()
        except Exception:
            pass


def get_lan_ips():
    ips = set()
    try:
        for info in socket.getaddrinfo(socket.gethostname(), None, socket.AF_INET):
            ip = info[4][0]
            if not ip.startswith("127."):
                ips.add(ip)
    except Exception:
        pass
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("10.255.255.255", 1))
        ips.add(s.getsockname()[0])
        s.close()
    except Exception:
        pass
    return sorted(ips) or ["127.0.0.1"]


def main():
    if not os.path.isdir(ROOT):
        print(f"[错误] 共享目录不存在: {ROOT}")
        sys.exit(1)
    data = load_auth()
    httpd = ThreadingHTTPServer((HOST, PORT), Handler)
    ips = get_lan_ips()
    print("=" * 60)
    print("  技术部项目资料数据库 已启动 ✅ (v4)")
    print(f"  本机访问:  http://localhost:{PORT}")
    for ip in ips:
        print(f"  同事访问:  http://{ip}:{PORT}")
    print(f"  数据根目录:  {ROOT}")
    print(f"  账号: {len(data['accounts'])} 个 | 待审批: "
          f"{sum(1 for a in data['applications'] if a['status']=='pending')} 条")
    print("  按 Ctrl+C 停止服务")
    print("=" * 60)
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        print("\n服务已停止")
        httpd.server_close()


if __name__ == "__main__":
    main()
